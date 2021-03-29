import os
import configparser
import shutil
import re
from openpyxl import load_workbook
import pandas as pd
from Plotter import Plot
from inspec_Logger import get_log

# read config file
config = configparser.RawConfigParser()
config.read("../cfg/cfg.ini")
TEMP_path = config.get("path", "temp_path_crawler")
TARGET_path = config.get("path", "target_path")
XLSX_path = config.get("path", "name_xlsx")
logger = get_log()

type_list = [".py"]
file_dict = {}
output_dict = {}
scope_dict = {}

Find = False
Total_Files = 0
Total_Lines = 0
Total_Braches = 0
Empty_Branches = 0
No_Result_Branches = 0

# create folder if doesnt exist
if not os.path.exists(TEMP_path):
    os.mkdir(config.get("path", "temp_path"))


# find if current line of code has target method
# return its index
def find_index(target, current_line):
    tar, tar_cap, tar_upp, tar_low = 0, 0, 0, 0
    if target in current_line:
        tar = current_line.find(target)
    if cap(target) in current_line:
        tar_cap = current_line.find(cap(target))
    if target.upper() in current_line:
        tar_upp = current_line.find(target.upper())
    if target.lower() in current_line:
        tar_low = current_line.find(target.lower())

    return tar, tar_cap, tar_upp, tar_low


# output the scope if located
def output_scope(branch, scope_full_name, scopes_requirement, output_dict, zipped):
    file, count = str(zipped[0]), str(zipped[1])
    print("Found the scope from: "+file)
    print(scope_full_name + ": ")
    print(scopes_requirement.tolist())
    print()
    logger.info("["+branch +"] "+"find scope: "+scope_full_name + " => "+str(scopes_requirement.tolist())+" from: "+file+" line "+count)
    scope_list = output_dict[branch]
    requirement_list = scopes_requirement.tolist()

    if requirement_list != []:
        temp_dict = {scope_full_name: requirement_list[0]}
    else:
        temp_dict = {scope_full_name: []}
    found = False

    if scope_list == []:
        scope_list.append(temp_dict)
        output_dict[branch] = scope_list
    else:
        for scope in scope_list:
            if temp_dict == scope:
                found = True
        if not found:
            scope_list.append(temp_dict)
            output_dict[branch] = scope_list


# looking for scope
# this method will recursively parsing the current line of code
# and find all api it uses
def inspect_keyword(scope_dict):
    global Find, No_Result_Branches, Total_Braches, Empty_Branches, Total_Files, Total_Lines
    data = pd.read_excel(str(XLSX_path))
    scopes = list(scope_dict.keys())
    pages = os.listdir(TEMP_path)

    for page in pages:
        print("-------------------------------------------------------------page: " + page)
        branch_path = TEMP_path+page
        branches = os.listdir(branch_path)
        for branch in branches:

            temp_dict = {branch: []}
            output_dict.update(temp_dict)
            print("【branch】: " + branch)
            files_path = TEMP_path+page+"/"+branch
            files = os.listdir(files_path)
            if len(files) == 0:
                Empty_Branches += 1
            for file in files:
                Total_Files += 1
                Count_lines = 0
                for line in open(TEMP_path+page+"/"+branch+"/"+file, encoding="utf-8"):
                    Total_Lines += 1
                    Count_lines += 1
                    #  firstly search for the first subname of the method
                    for scope in scopes:
                        if (scope in line) or (cap(scope) in line)or (scope.upper() in line)or (scope.lower() in line):

                            sc, sc_cap, sc_upp, sc_low = find_index(scope, line)

                            scope_sorted_list = [sc, sc_cap, sc_upp, sc_low]
                            scope_sorted_list.sort(reverse=True)
                            scope_string_index = scope_sorted_list[0]
                            scope_string_end_index = scope_string_index + len(scope)

                            # once locate category, looking for method of it
                            sub_scopes = scope_dict[scope]
                            # split the line of code and take latter.
                            temp = line[scope_string_index:]

                            before_scope_string_end_index = scope_string_end_index-scope_string_index
                            # print("总权限名: " + str(before_scope_string_index))
                            # print("总权限名截至: " + str(before_scope_string_end_index))

                            # if locates the first subname, starting searching for the rest
                            for sub_scope in sub_scopes:
                                for sub_scope_part in sub_scope:
                                    # print("当前搜索权限名: "+sub_method)
                                    cur_temp, \
                                    cur_scope_string_index, \
                                    cur_scope_string_end_index \
                                        = inspect_sub_scope(branch, data, scope, sub_scope, sub_scope_part, temp,
                                                            before_scope_string_end_index, output_dict, [file, Count_lines])

                                    if cur_temp and cur_scope_string_index and cur_scope_string_end_index:
                                        # if all three vars are not None replace the previous vars by current var
                                        # in order to start another round of recursion
                                        before_scope_string_end_index = cur_scope_string_end_index
                                        temp = cur_temp
                                    else:
                                        # current method doesnt appear in this line of code, break the loop
                                        # searching for the next method
                                        break
            if not Find and len(files)!=0: No_Result_Branches+=1
            Total_Braches += 1
            Find = False

# component of method [inspection]
# the search is separated to two parts, one is searching for the first subname of the scope,
# the other is searching for the rest scopes.
# this method is the latter one.


def inspect_sub_scope(branch, data, scope, sub_scope, sub_scope_part, temp, before_scope_string_end_index, output_dict,zipped):
    global Find
    if (sub_scope_part in temp) or (cap(sub_scope_part) in temp) or (sub_scope_part.upper() in temp)or (sub_scope_part.lower() in temp):
        tp, tp_cap, tp_upp, tp_low = find_index(sub_scope_part, temp)

        method_sorted_list = [tp, tp_cap, tp_upp, tp_low]

        method_sorted_list.sort(reverse=True)
        method_string_index = method_sorted_list[0]

        sub_temp = temp[method_string_index:]
        size = len(sub_scope)
        # print("当前遍历权限列表: "+str(method))
        # print("切分后文本: " + sub_temp)
        # method_string_end_index = method_string_index + before_scope_string_index + len(sub_method)
        method_string_end_index = method_string_index + len(sub_scope_part)

        distance = (method_string_index - before_scope_string_end_index) >= 0 and (
                    method_string_index - before_scope_string_end_index) <= 1
        # print("size: "+str(size))
        # print("distance: "+ str(distance))
        if size == 1 and distance:
            scopes_requirement = data[data.name == str(scope + "." + sub_scope[0])].scope
            scope_full_name = scope + "." + sub_scope[0]
            output_scope(branch, scope_full_name, scopes_requirement, output_dict, zipped)
            Find = True
        elif size > 1 and distance:
            count = 1
            scope_full_name = scope + "." + sub_scope[0]
            for sub_method_name in sub_scope[1:]:
                # print("sub_method_name: "+sub_method_name)
                # print("sub_temp: "+sub_temp)
                if (sub_method_name in sub_temp) or (sub_method_name.capitalize() in sub_temp) or (
                        sub_method_name.upper() in sub_temp or sub_method_name.lower() in sub_temp):
                    scope_full_name = scope_full_name + "." + sub_method_name
                    count = count + 1
            if count == size:
                scopes_requirement = data[data.name == str(scope_full_name)].scope
                output_scope(branch, scope_full_name, scopes_requirement, output_dict, zipped)
                Find = True
            else:
                # if this method returns 3 None, which means it failed finding a sub method name of current method
                return None, None, None
        else:
            return None, None, None
        return sub_temp, method_string_index, method_string_end_index
    else:
        return None, None, None


# build a binary tree according to xlsx list
def scope_dict() -> dict:
    method_dict = {}
    workbook = load_workbook(str(XLSX_path))
    booksheet = workbook.active
    for i in range(2, booksheet.max_row+1):
        method = booksheet.cell(i, 1).value
        key = method.split(".")[0]
        if key in method_dict.keys():
            methods_list = list(method.split(".")[1:])
            methods_list = reconstruct_list(methods_list)
            method_dict[key].append(methods_list)
        else:
            methods_list = list(method.split(".")[1:])
            methods_list = reconstruct_list(methods_list)
            temp_dict = {str(key): [methods_list]}
            method_dict.update(temp_dict)
    return method_dict


def reconstruct_list(litarable_list):
    count = True
    for element_key in litarable_list:
        if count:
            words, valid = find_cap(element_key)
            if valid:
                remove_index = litarable_list.index(element_key)
                litarable_list.pop(remove_index)
                index = remove_index
                for word in words:
                    litarable_list.insert(index, word)
                    index = index + 1
                count = False

    return litarable_list


def copy(file_dict):
    # copy files
    for key, value in zip(file_dict.keys(), file_dict.values()):
        shutil.copy(value, TEMP_path+key)
    # print("Finished search and copy")


def search(path, file_dict):
    # search for .py(currently)files
    files = os.listdir(path)
    for file in files:
        file_path = os.path.join(path, file)
        if os.path.isdir(file_path):
            search(file_path, file_dict)
        else:
            #
            if not has_key(file, file_dict):
                continue
            else:
                temp_dict = {str(has_key(file, file_dict)): str(file_path)}
                file_dict.update(temp_dict)
    return file_dict


def has_key(key, file_dict):
    (filename, extension) = os.path.splitext(key)
    if not extension in type_list:
        return
    new_key = filename + ".txt"
    count = 0
    while new_key in file_dict.keys():
        count = count+1
        new_key = filename+str(count)+".txt"
    return new_key


def cap(text):
    text = text[0].upper()+text[1:]
    return text


def find_cap(text):
    # find if theres capital letter
    cap_num = re.findall(r'[A-Z]', text)
    if len(cap_num) > 0:
        first_cap_index = text.find(cap_num[0])
        lower_case_word = text[0:first_cap_index]
        caplist = re.findall(r'[A-Z](?:[A-Z]*(?![a-z])|[a-z]*)', text)
        caplist.insert(0, lower_case_word)
        return caplist, True
    return None, False


def search_copy(file_dict):
    # create folder if doesnt exist
    if not os.path.exists(TARGET_path):
        os.mkdir(config.get("path", "target_path"))

    dict = search(TARGET_path, file_dict)
    copy(dict)


if __name__ == "__main__":

    # this method is used to search for all python files in [target] folder (which will be created if doesnt exist)
    # after searching it will convert py file to text file in [temp] folder for further inspection
    # it is useless when inspecting code from crawler
    # search_copy(file_dict)

    logger.info("-------------------------------start parsing-------------------------------")
    # # scope list
    scope_dict = scope_dict()
    # # inspection method
    inspect_keyword(scope_dict)
    if XLSX_path == '../cfg/slackname.xlsx':
        print("Current API: Events API")
    else:
        print("Current API: WEB API")
    print("Parsed " + str(Total_Braches) + " branches in total.\n")
    print("There are " + str(Total_Files) + " files "+ str(Total_Lines)+" lines in total.\n")
    print("There are " + str(Empty_Branches) + " empty branches.\n")
    print("There are "+str(No_Result_Branches)+" branches are not been detected.\n")
    # plot the diagram
    Plotter = Plot()
    count_dict = Plotter.manipulate_data(output_dict)

    Plotter.plot(count_dict)
