import os
import configparser
import re
from openpyxl import load_workbook

'''
This file contains the esstential methods for code checker. 
They either parse the xlsx file or process the data.
author: ZHAN YICHENG 03/04/2021
'''

# read config file
config = configparser.RawConfigParser()
config.read("../cfg/cfg.ini")
TEMP_path = config.get("path", "temp_path_JavaScript")
XLSX_path = config.get("path", "name_xlsx")
type_list = [".py"]


# build a data tree according to xlsx list
def scope_dict() -> dict:
    method_dict = {}
    workbook = load_workbook(str(XLSX_path))
    booksheet = workbook.active
    for i in range(2, booksheet.max_row+1):
        method = booksheet.cell(i, 1).value
        key = method.split(".")[0]
        if key in method_dict.keys():
            methods_list = list(method.split(".")[1:])
            methods_list = reconstruct_list(methods_list)
            method_dict[key].append(methods_list)
        else:
            methods_list = list(method.split(".")[1:])
            methods_list = reconstruct_list(methods_list)
            temp_dict = {str(key): [methods_list]}
            method_dict.update(temp_dict)
    return method_dict

# This method rebuild the API methods. It will split the method name that contains Capital letter.
# For example: 'chat.postMessage' will be split to chat.post,Message.
# It does this to deal with various calling ways.
def reconstruct_list(litarable_list):
    count = True
    for element_key in litarable_list:
        if count:
            words, valid = find_cap(element_key)
            if valid:
                remove_index = litarable_list.index(element_key)
                litarable_list.pop(remove_index)
                index = remove_index
                for word in words:
                    litarable_list.insert(index, word)
                    index = index + 1
                count = False

    return litarable_list

# if a repository contains multiple files with the same name, add extra prefix to avoid errors.
def has_key(key, file_dict):
    (filename, extension) = os.path.splitext(key)
    if not extension in type_list:
        return
    new_key = filename + ".txt"
    count = 0
    while new_key in file_dict.keys():
        count = count+1
        new_key = filename+str(count)+".txt"
    return new_key


# find if theres capital letter
def find_cap(text):
    cap_num = re.findall(r'[A-Z]', text)
    if len(cap_num) > 0:
        first_cap_index = text.find(cap_num[0])
        lower_case_word = text[0:first_cap_index]
        caplist = re.findall(r'[A-Z](?:[A-Z]*(?![a-z])|[a-z]*)', text)
        caplist.insert(0, lower_case_word)
        return caplist, True
    return None, False


# check if current line of code has target method
def find_index(target, current_line):
    tar, tar_cap, tar_upp, tar_low = 0, 0, 0, 0
    if target in current_line:
        tar = current_line.find(target)
    if cap(target) in current_line:
        tar_cap = current_line.find(cap(target))
    if target.upper() in current_line:
        tar_upp = current_line.find(target.upper())
    if target.lower() in current_line:
        tar_low = current_line.find(target.lower())
    # return the index of 4 data types
    return tar, tar_cap, tar_upp, tar_low


# Capitalize the first letter
def cap(text):
    text = text[0].upper()+text[1:]
    return text